import os
import shutil
import time
from datetime import datetime
from threading import Thread

import xlrd
from openpyxl import load_workbook

root_dir = r"D:\请上传学生信息相关Excel文档"
# serial_number = 'XME2019120800000'
serial_numberNo = 'NO.'

def func(n):
    time.sleep(3)
    print(n)

def trim(s):
    if len(s)==0:
        return ''
    if s[:1]==' ':
        return trim(s[1:])
    elif s[-1:]=='':
        return trim(s[:-1])
    else:
        return s

def removeFiles(top):
    for root, dirs, files in os.walk(top, topdown=False):
        print(root,dirs,files)
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.rmdir(os.path.join(root, name))
        print("正在删除文件：",root)
        os.rmdir(root)
    return 1

if __name__ == '__main__':
    s = input("请输入一个要解析的地址:")
    s = trim(s)
    try:
        completem_path = os.path.join(root_dir,s)
        print(completem_path)
        if not os.path.exists(completem_path):
            print("文件不存在")
            # time.sleep( 3 )
            t = Thread(target=func,args=("文件不存在",))
            t.start()
        else:
            book = xlrd.open_workbook(completem_path)
            # book = xlrd.open_workbook(os.path.join(root_dir, r"XMETC 诚毅学院就业实训295人结业证书登记总表-2019.12.8.xlsx"))
            # name_sheet = 'XMETC 软工3班就业实训40人结业证书登记表'
            ss = 0
            for name_sheet in book.sheet_names()[1:]:
                if   os.path.exists(os.path.join(root_dir, name_sheet)):
                    # print(os.path.join(root_dir, "temp"))
                    shutil.move(os.path.join(root_dir, name_sheet), os.path.join(root_dir, "temp"))  #复制一个文件到一个文件或一个目录
                    # shutil.os.rmdir(os.path.join(root_dir, name_sheet))
                    shutil.rmtree( os.path.join(root_dir, "temp"))

            for name_sheet in book.sheet_names()[1:]:
                if not os.path.exists(os.path.join(root_dir, name_sheet)):
                    os.makedirs(os.path.join(root_dir, name_sheet))
                    print(os.path.join(root_dir, name_sheet))
                else:
                    shutil.os.rmdir(os.path.join(root_dir, name_sheet))
                    os.makedirs(os.path.join(root_dir, name_sheet))

                sheet = book.sheet_by_name(name_sheet)
                wb = load_workbook(os.path.join(root_dir, r'就业证书-横版.xlsx'))
                wb_vertical = load_workbook(os.path.join(root_dir, r'实训证书-竖版.xlsx'))
                sheet_ranges = wb['横版-正面']
                sheet_ranges_vertical = wb_vertical['竖版-正面']

                serial_number_across = load_workbook(os.path.join(root_dir, r'就业证书编号-横版2.xlsx'))
                serial_number_vertical = load_workbook(os.path.join(root_dir, r'实训证书编号-竖版2.xlsx'))
                sheet_serial_number_across = serial_number_across['Sheet1'] # 横版的excel下面的表名
                sheet_serial_number_vertical = serial_number_vertical['Sheet1'] # 竖版版的excel下面的表名
                for i in range(5, sheet.nrows):
                    row = sheet.row_values(i)
                    row[5] = (row[5] - 19 - 70 * 365) * 86400 - 8 * 3600
                    row[6] = (row[6] - 19 - 70 * 365) * 86400 - 8 * 3600
                    row[5] = datetime.fromtimestamp(row[5])
                    row[6] = datetime.fromtimestamp(row[6])
                    # 需要读取彪哥中的单元格数据的证书编号
                    certificate_number = row[1]
                    print("读取编号： ",certificate_number)
                    # serial_number_temp = serial_number[:-len(str(int(row[1])))] + str(int(row[1]))
                    sheet_serial_number_across['B3'] = serial_numberNo
                    sheet_serial_number_across['C3'] = certificate_number
                    serial_number_across.save(os.path.join(root_dir, f'{name_sheet}\就业证书编号-横版{int(row[0])}-{row[7]}.xlsx'))
                    sheet_serial_number_vertical['B3'] = serial_numberNo
                    sheet_serial_number_vertical['C3'] = certificate_number
                    serial_number_vertical.save(os.path.join(root_dir, f'{name_sheet}\实训证书编号-竖版{int(row[0])}-{row[7]}.xlsx'))

                    sheet_ranges['C18'] = row[7]
                    sheet_ranges['E18'] = row[5].year
                    sheet_ranges['F18'] = row[5].month
                    sheet_ranges['G18'] = row[5].day
                    sheet_ranges['I18'] = row[6].year
                    sheet_ranges['J18'] = row[6].month
                    sheet_ranges['K18'] = row[6].day
                    sheet_ranges['A20'] = '                      ' + row[3]
                    wb.save(os.path.join(root_dir, f'{name_sheet}\就业证书-横版{int(row[0])}-{row[7]}.xlsx'))

                    sheet_ranges_vertical['C14'] = row[7]
                    sheet_ranges_vertical['E14'] = row[5].year
                    sheet_ranges_vertical['F14'] = row[5].month
                    sheet_ranges_vertical['G14'] = row[5].day
                    sheet_ranges_vertical['I14'] = row[6].year
                    sheet_ranges_vertical['J14'] = row[6].month
                    sheet_ranges_vertical['K14'] = row[6].day
                    sheet_ranges_vertical['F16'] = '       ' + row[3]
                    wb_vertical.save(os.path.join(root_dir, f'{name_sheet}\实训证书-竖版{int(row[0])}-{row[7]}.xlsx'))
                ss = ss +1
                if ss == 3:
                    ss = ""

    except Exception as err:
        print(err)
        input('程序错误，请按下任意键结束进程...')


